import numpy
import scipy.special
import astropy.io.fits as pyfits
import os
import numpy as np
import glob
import openpyxl
import lmfit
from importlib import reload
import fit_ionrec
reload(fit_ionrec)
from fit_ionrec import younger, mewe
import ipdb
import pickle
import matplotlib
#matplotlib.use('TkAgg')

# functions to calculate the ionization rate coefficients due to
# direct collisional ionization (ci) and excitation-autoionization (ea)

# see example at the end

# I haven't made the function for cross sections, which you will need
# to compare the your experiments. 
# 
# Copy them from Urdampilleta, Kaastra & Mehdipour 2017 
# http://www.aanda.org/10.1051/0004-6361/201630170

# example is at the end


KBOLTZ = 8.617385e-8
ME_KEV = 510.99895
APEDDIR = 'APED' # THIS IS THE DIRECTORY WHERE YOU HAVE PUT THE IR FILES
DATADIR = 'csdata'# THIS IS THE DIRECTORY WHERE YOU HAVE PUT THE CROSS SECTION DATA FILES
def calc_ci_urdam(T, param):
  """
  Calculate the collisional electron impact ionization rates using the
  Urdampilletta formulae, for 1 shell

  Parameters
  ----------
  Te : float or array(float)
    Electron temperature (K)
  ionpot : float
    Ionization potential (eV)
  param : array(float)
    The parameters of the equation
    shell no, Ionization potential (keV), 8 parameters

  Returns
  -------
  float or array(float)
    Ionization rate in cm^3 s^-1

  References
  ----------
  Urdampilletta
  """
  import scipy.special
  g7a1 = 3.480230906913262     # sqrt(pi)*(gamma + ln(4))
  g7a2 = 1.772453850905516     # sqrt(pi)
  g7a3 = 1.360544217687075E-02 # 2/147
  g7a4 = 0.4444444444444444    # 4/9
  g7c1 = -4881/8.
  g7c2 =  1689/16.
  p7 = numpy.array([1.000224,-0.113011,1.851039,0.0197311,0.921832,2.651957])
  gamma = 0.577215664901532860606512    #Euler's constant
  a = numpy.array([0.999610841,3.50020361,-0.247885719,0.0100539168,1.39075390e-3,1.84193516,4.64044905])
  r0 = 4.783995473666830e-10   #=2*sqrt(2/pi)*c*1e-18

  ishell = int(param[0])
  eion = param[1] # in keV
  kT = T * KBOLTZ
  #  print('eion', eion)
  y = eion/kT
  lam = eion/ME_KEV

  # filter out when y > 200
  yhi = y>200.
  ylo = y<=200.
  en1 = numpy.zeros(len(y), dtype=float)

  en1[ylo] = numpy.exp(y[ylo]) * scipy.special.exp1(y[ylo])
  en1[yhi] = 0.001



  g=numpy.zeros((8, len(kT)))
  #print(y, numpy.exp(y))
  g[0,:] = 1/y
  g[1,:] = en1
  g[2,ylo] = scipy.special.expn(2, y[ylo])*numpy.exp(y[ylo])
  g[2,yhi] = 0.001

  g[3,:] = en1/y
  g[4,:] = (1+en1)/y**2
  g[5,:] = (3+y+2*en1)/y**3

  k = numpy.where(y<0.6)[0]
  if len(k) > 0:
    yy = y[k]
    g[6,k] = numpy.exp(yy) * \
          ((((yy / 486.0 - g7a3) * yy + 0.08) * yy - g7a4) * yy + 4 \
          - (g7a2 * numpy.log(yy) + g7a1) / numpy.sqrt(yy))

  k = numpy.where((y>=0.6) & (y<=20.0))[0]
  if len(k) > 0:
    yy=y[k]
    g[6,k] = (p7[0] + (p7[1] + p7[3] * numpy.log(yy)) /\
              numpy.sqrt(yy) + p7[2] / yy) / (yy + p7[4]) / (yy + p7[5])

  k = numpy.where(y>20.0)[0]
  if len(k) > 0:
    yy=y[k]
    g[6,k] = (((((g7c1 / yy + g7c2) / yy - 22.0) /\
                 yy + 5.75) / yy - 2) / yy + 1)/yy**2


  k = numpy.where(y<0.5)[0]
  if len(k) > 0:
    yy = y[k]
    g[7,k] = (((((-yy / 3000.0 - (1.0 / 384.0)) * yy - (1.0 / 54.0)) * yy +\
             0.125) * yy - 1.) * yy + 0.989056 + \
              (numpy.log(yy) / 2.0 + gamma) * numpy.log(yy)) * numpy.exp(yy)


  k = numpy.where(((y>=0.5) & (y<=20.0)))[0]
  if len(k) > 0:
    yy = y[k]
  #    print(yy)
    g[7,k] = ((((a[4] / yy + a[3]) / yy + a[2]) / yy + a[1]) / yy + a[0]) /\
               (yy + a[5]) / (yy + a[6])
  k = numpy.where(y>20.0)[0]
  if len(k) > 0:
    yy = y[k]
    g[7,k] = ((((((13068 / yy - 1764) / yy + 274) / yy - 50) / yy +\
               11) / yy - 3) / yy + 1) / yy**2

  ciout = param[2] * 1e24*(g[0,:] - g[1,:]) + \
          param[3] * 1e24* (g[0,:] - 2 * g[1,:] + g[2,:]) + \
          param[4] * 1e24* (g[3,:] + 1.5 * lam * g[4,:] + 0.25 * lam**2 * g[5,:]) + \
          param[5] * 1e24* g[6,:] +\
          param[6] * 1e24* g[7,:]
  ciout *= r0 * numpy.exp(-y) / eion**2 * y**1.5 * numpy.sqrt(lam)
  ciout[yhi]=0.0
  return(ciout)

def calc_ci_crosssec_urdam(elecE, param):
  """
  Calculate the collisional electron impact ionization cross sections using the
  Urdampilletta formulae, for 1 shell

  Parameters
  ----------
  elecE : float or array(float)
    Electron energy (keV)
  param : array(float)
    The parameters of the equation
    shell no, Ionization potential (keV), 8 parameters

  Returns
  -------
  float or array(float)
    Ionization cross section

  References
  ----------
  Urdampilletta
  """
  import scipy.special
  ishell = int(param[0])
  eion = param[1] # in keV

  #The formula looks like u*I^2*Q = A(1-1/u) + B(1-1/u)^2 + C*R*ln(u) + D*ln(u)/sqrt(u) + E*ln(u)/u (equation 2)
  # where u is electron energy E divided by the shell's ionization potential I (i.e., u=E/I)
  # Q is the cross section
  # A,B,C,D,E are the fitted parameters passed in with param
  # R is the relativistic correction term
  u = elecE/eion
  A = param[2]
  B = param[3]
  C = param[4]
  D = param[5]
  E = param[6]

  epsilon = elecE/ME_KEV
  tau = eion/ME_KEV
  R = ((tau+2)/(epsilon+2))* ( (epsilon+1)/(tau+1))**2 * \
       ( ((tau+epsilon)*(epsilon+2) *(tau+1)**2)/ (epsilon * (epsilon+2) * (tau+1)**2+tau*(tau+2)))**1.5

  #epsilon = elecE/(ME_KEV*(SPEED_OF_LIGHT**2))
  #R = 1+1.5*epsilon + 0.25*(epsilon**2)

  Aterm = A*(1-(1/u))
  Bterm = B*(1-(1/u))**2
  Cterm = C* R * np.log(u)
  Dterm = D* np.log(u)/np.sqrt(u)
  Eterm = E* np.log(u)/u

  ci_cross = 1/(u*eion*eion)*(Aterm + Bterm + Cterm + Dterm + Eterm)
  ci_cross[u<1] = 0.0
  return ci_cross * 10**(2*2) #in units of cm^-2

def calc_ci_sigma_urdam(E, param):
  """
  Calculate the collisional electron impact ionization rates using the
  Urdampilletta formulae, for 1 shell

  Parameters
  ----------
  E : float or array(float)
    Electron temperature (keV)
  param : array(float)
    The parameters of the equation
    shell no, Ionization potential (keV), 8 parameters

  Returns
  -------
  float or array(float)
    Cross section in cm^2

  References
  ----------
  Urdampilletta
  """
  import scipy.special
  
  I = param[1] # ionization parameter in keV
  A = param[2]
  B = param[3]
  C = param[4]
  D = param[5]
  F = param[6] # this is parameter E in the formula, but I don't want to confuse with "E" for Energy.
  
  tau = I/ME_KEV
  eps = E/ME_KEV
  R = ((tau+2)/(eps+2))* ( (eps+1)/(tau+1))**2 * \
       ( ((tau+eps)*(eps+2) *(tau+1)**2)/ (eps * (eps+2) * (tau+1)**2+tau*(tau+2)))**1.5

  u = E/I
  
  Q = A * (1- (1/u)) + B*(1-(1/u))**2 + C * R * numpy.log(u) + D * numpy.log(u)/numpy.sqrt(u) + F * numpy.log(u)/u
  
  Q = Q/(u*I*I)*10000 # m^2 -> cm^2
  
  # having done all that: at all points, if E < ionization potential (so u < 1), the cross section is 0.
  
  Q[u<1] = 0.0
  return(Q)
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------
#-------------------------------------------------------------------------------

def calc_ea_urdam(T, param):
  """
  Calculate the collisional excitation-autoionization rates using the
  Urdampilletta formulae, for 1 shell

  Parameters
  ----------
  Te : float or array(float)
    Electron temperature (K)
  param : array(float)
    The parameters of the equation
    shell no, Ionization potential (keV), 8 parameters

  Returns
  -------
  float or array(float)
    Ionization rate in cm^3 s^-1

  References
  ----------
  Urdampilletta
  """
  import scipy.special
  r0 = 4.783995473666830e-10   #=2*sqrt(2/pi)*c*1e-18

  ishell = int(param[0])
  eion = param[1] # in keV
  kT = T * KBOLTZ
  y = eion/kT
  lam = eion/ME_KEV
  exp1=scipy.special.exp1(y)
  # filter out when y > 200
  yhi = y>200.
  ylo = y<=200.
  #en1 = numpy.zeros(len(y), dtype=float)

  #en1[ylo] = numpy.exp(y[ylo]) * exp1[ylo]
  #en1[yhi] = 0.001

  #eminusy = numpy.exp(-y)

  #m1 = (1/y)* eminusy
  #m2 = exp1
  #m3 = eminusy- y*exp1
  #m4 = (1-y)*eminusy/2 + y**2*exp1/2
  #m5 = exp1/y

  em1 = numpy.zeros(len(y))
  em2 = numpy.zeros(len(y))
  em3 = numpy.zeros(len(y))

  em1[ylo] = scipy.special.expn(1, y[ylo])*numpy.exp(y[ylo])
  em2[ylo] = scipy.special.expn(2, y[ylo])*numpy.exp(y[ylo])
  em3[ylo] = scipy.special.expn(3, y[ylo])*numpy.exp(y[ylo])

  c_ea = param[2]*1e24 + \
         y * (param[3]*1e24*em1 + param[4]*1e24 * em2 + 2*param[5]*1e24 * em3) + param[6]*1e24*em1

  c_ea *=  r0*numpy.exp(-y)/eion**2 * y**1.5 * lam**0.5
  c_ea[yhi]=0.0


  return(c_ea)

def calc_ea_crosssec_urdam(elecE, param):
  """
  Calculate the collisional excitation-autoionization cross sections using the
  Urdampilletta formulae, for 1 shell

  Parameters
  ----------
  elecE : float or array(float)
    Electron energy (keV)
  param : array(float)
    The parameters of the equation
    shell no, Ionization potential (keV), 8 parameters

  Returns
  -------
  float or array(float)
    Ionization cross section

  References
  ----------
  Urdampilletta
  """
  import scipy.special
  ishell = int(param[0])
  eion = param[1] # in keV

  #The formula looks like u*I^2*Q = A + B/u + C/(u**2) + 2*D/(u**3) + E*ln(u) (equation 12)
  # where u is electron energy E divided by the shell's ionization potential I (i.e., u=E/I)
  # Q is the cross section
  # A,B,C,D,E are the fitted parameters passed in with param
  # R is the relativistic correction term
  u = elecE/eion
  A = param[2]
  B = param[3]
  C = param[4]
  D = param[5]
  E = param[6]

  Aterm = A
  Bterm = B/u
  Cterm = C/(u**2)
  Dterm = 2*D/(u**3)
  Eterm = E*np.log(u)

  ea_cross = 1/(u*(eion**2))*(Aterm + Bterm + Cterm + Dterm + Eterm)# in units of 10^-24 m^2
  ea_cross[u<1] = 0.0
  return ea_cross * 10**(2*2) #in units of cm^-2

def get_params(Z, z1, elsymb):
  # Z = element nuclear charge
  # z1 = ionizing ion charge +1 
  # elsymb  = element atomic symbol (e.g. Be, Fe etc)
  
  # the parameters are, for each shell:
  #    the index (so just 1, 2, 3,...),
  #    the ionization potential of the shell (in eV)
  #    the A, B, C, D, E parameters from Urdampilleta.
  
  # Both exciation-autoionization and CI use the same number of and name for 
  # the parameters, but they use different formulae to convert to
  # cross sections/rates
  
  # The bethe limit it parameter C, of CI from shell 1.

  #todo: add c
  
  
  # get the filename

  searchstring = os.path.abspath('%s/%s/%s_%i/%s_%i_IR*fits'%(APEDDIR,
                                             elsymb.lower(),
                                             elsymb.lower(),
                                             z1,
                                             elsymb.lower(),
                                             z1))
  
  fname = glob.glob(searchstring)
  #print(searchstring, fname)
  irdat = pyfits.open(fname[0])
  
  # get the data:
  
  icidat = numpy.where(irdat[1].data['PAR_TYPE']==52)[0]
  
  cidat = []
  for i in icidat:
    cidat.append(numpy.array(irdat[1].data['IONREC_PAR'][i][:7], dtype=float))

    
  ieadat = numpy.where(irdat[1].data['PAR_TYPE']==64)[0]
  
  eadat = []
  for i in ieadat:
    eadat.append(numpy.array(irdat[1].data['IONREC_PAR'][i][:7], dtype=float))
  

  
  ret={'ci':cidat,
       'ea':eadat}

  return(ret)


def get_ionrec_rate(Z, z1, elsymb, T):
  
  # get the atomic data
  
  dat = get_params(Z, z1, elsymb) #gets ABCDE from atomdb/Urdampilleta
  
  eaout = numpy.zeros(len(T))
  ciout = numpy.zeros(len(T))
  for ea in dat['ea']:
    eaout += calc_ea_urdam(T, ea )
    
  for ci in dat['ci']:
    tmp = calc_ci_urdam(T, ci)
    ciout += tmp
  return(ciout, eaout)
    
def get_ionrec_crosssec(Z, z1, elsymb, elecE):
  
  # get the atomic data
  
  dat = get_params(Z, z1, elsymb) #gets ABCDE from atomdb/Urdampilleta
  
  eaout = numpy.zeros(len(elecE), dtype=np.float64)
  ciout = numpy.zeros(len(elecE), dtype=np.float64)
  for ea in dat['ea']:
    eaout += np.array(calc_ea_crosssec_urdam(elecE, ea), dtype=np.float64)
  for ci in dat['ci']:
    tmp = np.array(calc_ci_crosssec_urdam(elecE, ci), dtype=np.float64)
    ciout += tmp
    
  return(ciout, eaout)    

def ionrecAnalysis(Z=5, finalChargeState=3, elementSymbol='B', monteCarloLength=100, numCIModels=1, numEAModels=1, 
                   lowTempPower=2, highTempPower=15, numTempSteps=60, makePlots=False, makePlotsRates=True, FixD=True, addZeroPt=False):
  """
  Calculate the collisional excitation-autoionization cross sections using the
  Urdampilletta formulae and compare it to experimental data. Get rates and uncertainties from fitting
  experimental data with a Monte Carlo simulation.

  Parameters
  ----------
  Z : int
    Atomic number
  finalChargeState : int
    Charge of the ion after ionization
  elementSymbol : str
    Elemental symbol corresponding to Z, e.g. 'B' for Z=5
  monteCarloLength : int
    Number of Monte Carlo runs to get the average rates and uncertainties
  numCIModels : int
    Number of collisional ionization curves to include in the model according to the Younger formula. 
    There cannot be more than the number of CI models in AtomDB.
    Using fewer models (~1) makes sense for our low number of data points.
  numEAModels : int
    Number of excitation-autoionization curves to include in the model according to the Mewe formula. 
    There cannot be more than the number of EA models in AtomDB.
    Using fewer models (~1) makes sense for our low number of data points.
  lowTempPower : int
    Lower limit to use for calculating ionization rates, in K, as a power of 10.
    e.g., lowTempPower=2 --> rates start at 10^2 K
  highTempPower : int
    Upper limit to use for calculating ionization rates, in K, as a power of 10.
    e.g., highTempPower=15 --> rates end at 10^15 K
  numTempSteps : int
    How many temperatures to calculate rates for between temperatures specified by lowTempPower and highTempPower, in log space.
  makePlots : bool
    Make plots for the raw data fitting and monte carlo fitting. Warning: makes lots of plots--set monteCarloLength to <=3
  makePlotsRates : bool
    Make plots for rates after Monte Carlo simulation showing the mean and standard deviation.
  FixD : bool
    Whether or not fix the "D" parameter of each CI curve to 0, as mentioned in Urd. Good for stopping overfitting.
  addZeroPt : bool
    Optionally add a data point -- cross section=0 at the lowest ionization energy. Useful if the data set doesn't have low energy data points.
  Returns
  -------
  Ionization rates and uncertainties at temperature points : dict[list, list, list]
    Has the format {"Temperature (K)":Tlist, "Average exp. rates":avgExpRates, "1 sigma uncertainty":avgExpRatesUnc}
  """
  z1 = finalChargeState #final charge state
  elsymb = elementSymbol

  #load experimental values for plotting
  #searchstringexp = os.path.relpath('%s/%s/%s\%s%i+*xlsx'%(DATADIR,elsymb.upper(),"SI",elsymb.upper(),z1-1))
  searchstringexp = os.path.join(os.path.dirname(__file__), DATADIR,elsymb.upper(),"SI",f'{elsymb.upper()}{int(z1-1)}+.xlsx')
  fname = glob.glob(searchstringexp)[0]
  #print(searchstringexp, fname)

  if fname:
    #load energies, data
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    ws_contents=list(ws.iter_cols(values_only=True))
    for col in ws_contents:
      if col[0] == None:
        continue
      if "Energy" in col[0]:
        Elist = []
        tempE = np.array(col[1:], dtype=float)
        for E in tempE:
          if not np.isnan(E):
            Elist.append(E)
      elif col[0] == "cross section":
        csData = []
        tempcs = np.array(col[1:], dtype=float)
        for cs in tempcs:
          if not np.isnan(cs):
            csData.append(cs)
      elif col[0] == "cross section uncertainty":
        csUnc = []
        tempUnc = np.array(col[1:], dtype=float)
        for unc in tempUnc:
          if not np.isnan(unc):
            csUnc.append(unc)

      elif col[0] == "cross section power (cm2)":
        csPower = col[1]

    Elist = np.array(Elist)
    csData = np.array(csData)
    csUnc=np.array(csUnc)
    Elist=Elist[Elist!=None]*(10**-3) # in keV
    csData=csData[csData!=None]*(10**csPower)
    csUnc =csUnc[csUnc!=None]*(10**csPower)
  else:
    Elist = 2*np.logspace(1,3,num=20)*(10**-3) #in keV

  #ci means collisional electron impact (i.e., direct) ionization rate
  #ea means excitation-autoionization rate
  ci, ea = get_ionrec_crosssec(Z, z1, elsymb, Elist)
  # find the Bethe limit (C)
  params = get_params(Z, z1, elsymb)
  for i in range(len(params['ci'])):
    if int(params['ci'][i][0]) == 1:
       C = params['ci'][i][4]
       break
  if elementSymbol=='O' and finalChargeState==1: #remove extra CI curve in neutral oxygen from atomdb
    params['ci']=params['ci'][:-1]

  if makePlots:
    fig = plt.figure()
    fig.show()
    ax = fig.add_subplot(111)
    ax.semilogy(Elist, ci, label='collisional ionization')
    ax.semilogy(Elist, ea, label='excitation-autoionization')
    ax.semilogy(Elist, ci+ea, label='Total cross section', marker='o')
    if fname:
      ax.errorbar(Elist, csData, csUnc, label="experiment", marker='o', linestyle="None")
      #plt.xscale('log')
      plt.yscale('log')
    ax.legend(loc=0)
    plt.title(f"cross section of {elsymb} {z1-1}+ to {elsymb} {z1}+")
    plt.ylabel('Cross section (cm-2)')
    plt.xlabel('Electron energy (keV)')
    #plt.draw()
  #ipdb.set_trace()

  # numCIModels = 1#len(params['ci'])
  # numEAModels = 1#len(params['ea'])
  ###limit number of curves based on the number of available data points?
  maxModels = np.floor(len(Elist)/5)
  numCIModels = int(min(numCIModels, np.floor(maxModels/2), len(params['ci'])))
  numEAModels = int(min(numEAModels, int(maxModels-numCIModels), len(params['ea'])))
  lastInd = len(params['ci'])

  paramsDict = {} #will look like {1:{'ea1_eion':x, 'ea1_A':A, ...}, 2:{...}, N:{...}} where N is the number of monte carlo runs
  paramsPrefixes = []
  ElistOrig=Elist
  #repeat monteCarloLength times:
  for mcInd in range(1,monteCarloLength+1):
    csDataMonte = np.random.normal(loc=csData, scale=csUnc)
    crossSecModel = lmfit.Model(younger, prefix=f'ci{lastInd}_') #assume there's always a CI curve
    crossSecParams = crossSecModel.make_params(**dict(zip([f'ci{lastInd}_eion',f'ci{lastInd}_A',f'ci{lastInd}_B',f'ci{lastInd}_C',f'ci{lastInd}_D',f'ci{lastInd}_E'], params['ci'][-1][1:])))
    crossSecParams[f'ci{lastInd}_eion'].set(vary=False) # fix ionization potential
    crossSecParams[f'ci{lastInd}_C'].set(vary=False) # fix ionization potential
    if FixD: #urdam: D is sometimes set to 0
      crossSecParams[f'ci{lastInd}_D'].set(value=0,vary=False) # fix ionization potential
    CIprefixes = [f'ci{lastInd}_']
    EAprefixes = []
    LowestEion = params['ci'][-1][1]

    for i in range(numCIModels-1):
      if i<lastInd:
        ciInd = lastInd-(i+1)
        prefix = f'ci{ciInd}_'
        tempModel = lmfit.Model(younger, prefix=prefix)
        tempParams = tempModel.make_params(**dict(zip([f'{prefix}eion',f'{prefix}A',f'{prefix}B',f'{prefix}C',
                                                      f'{prefix}D',f'{prefix}E'], params['ci'][ciInd-1][1:])))
        CoeffMax = 1.e-3 #gets close to rates with limit as 10e-24 but doesnt follow cross sections
        CoeffMin = -1.e-3
        tempParams[f'{prefix}A'].set(max=CoeffMax, min=CoeffMin)
        tempParams[f'{prefix}B'].set(max=CoeffMax, min=CoeffMin)
        tempParams[f'{prefix}C'].set(max=CoeffMax, min=CoeffMin)
        tempParams[f'{prefix}D'].set(max=CoeffMax, min=CoeffMin)
        tempParams[f'{prefix}E'].set(max=CoeffMax, min=CoeffMin)

        tempParams[f'{prefix}C'].set(vary=False) #vary=False will fix C to the bethe limit
        if FixD:
          tempParams[f'{prefix}D'].set(value=0, vary=False) #vary=False will fix D
        tempParams[f'{prefix}eion'].set(vary=False) # fix ionization potential

        CIprefixes.append(prefix)

        crossSecModel += tempModel
        crossSecParams += tempParams
        if params['ci'][ciInd][1] < LowestEion:
          LowestEion = params['ci'][ciInd][1]

    for i in range(numEAModels):
      prefix = f'ea{i+1}_'
      tempModel = lmfit.Model(mewe, prefix=prefix)
      tempParams = tempModel.make_params(**dict(zip([f'{prefix}eion',f'{prefix}A',f'{prefix}B',f'{prefix}C',
                                                      f'{prefix}D',f'{prefix}E'], params['ea'][i][1:])))
      #tempParams[f'{prefix}C'].set(vary=False) #vary=False will fix C to the bethe limit
      tempParams[f'{prefix}eion'].set(vary=False) # fix ionization potential
                              
      EAprefixes.append(prefix)

      crossSecModel += tempModel
      crossSecParams += tempParams

      if params['ea'][i][1] < LowestEion:
        LowestEion = params['ea'][i][1]

    paramsPrefixes = EAprefixes+CIprefixes
    if addZeroPt:
      Elist=np.append([LowestEion],ElistOrig)
      csDataMonte=np.append([0],csDataMonte)
    def objective_custom(params, data, x):
      moreXvalues = np.logspace(np.log10(x[0]),np.log10(x[-1]),num=200)
      residual = data-crossSecModel.eval(params, electronEnergy=x)
      paramsValuesDict = crossSecModel.eval_components(**params, electronEnergy=moreXvalues)
      #negativePenalty=np.zeros(len(paramsValuesDict)*len(data))
      negativePenalty = np.array([])
      derivPenalty = np.array([])
      for key, comp in paramsValuesDict.items():
        negativePenalty = np.concatenate([negativePenalty,[abs(val *1000) if val < 0 else 0 for val in comp]],0)
        # derivPenalty = np.concatenate([derivPenalty,[(comp[i+1] - comp[i]) ** 2
        #                                              if i < len(comp) - 1 else 0
        #                                              for i in range(len(comp))]],0)
        if ('ea' in key):
          derivPenalty = np.concatenate([derivPenalty,[abs(comp[i+1] - 2 * comp[i] + comp[i-1])*1000 #does this penalize large ea curves rather than the curvature? exp changes but with a small step size...
                                                if i > 0 and i < len(comp) - 1 and comp[i-1]!=0 else 0
                                                for i in range(len(comp))]],0)
          
          # if any(comp<0):
          #   negativePenalty += 1000000000000 #very large residual for any negative values in each component
      return np.concatenate([residual,negativePenalty, derivPenalty],0)
    
    #result = crossSecModel.fit(csDataMonte[Elist>LowestEion], params=crossSecParams, electronEnergy=Elist[Elist>LowestEion])
    resultMinimizer = lmfit.minimize(objective_custom, crossSecParams, args=(csDataMonte[Elist>LowestEion], Elist[Elist>LowestEion]))
    # ipdb.set_trace()
    numPoints = 100
    xs = np.logspace(np.log10(min(Elist)), np.log10(max(Elist)), num=numPoints)
    if makePlots: 
      plt.figure()
      #result.plot(numpoints=numPoints)
      resultdata = crossSecModel.eval(params=resultMinimizer.params, electronEnergy=xs)
      plt.plot(Elist, csDataMonte, linestyle="None", marker="o", label="Data")
      plt.plot(xs, resultdata, label='Best fit')
      plt.ylim(np.min(csData[csData>0])/10, np.max(csData)*10)
    
    # for prefix, component in result.eval_components().items():
    #   plt.plot(Elist[Elist>LowestEion], component, '--', label=prefix)
    ### plot fitted components
    #bv = result.best_values
    bv = resultMinimizer.params.valuesdict()

    if makePlots:
      for prefix in CIprefixes:
        cidata = younger(xs, **dict(zip(['eion','A','B','C','D','E'], 
                                        [bv[f'{prefix}eion'],bv[f'{prefix}A'],bv[f'{prefix}B'],
                                        bv[f'{prefix}C'],bv[f'{prefix}D'],bv[f'{prefix}E']])))
        plt.plot(xs, cidata, '--',label=prefix, linewidth=1)
      for prefix in EAprefixes:
        eadata = mewe(xs, **dict(zip(['eion','A','B','C','D','E'], 
                                        [bv[f'{prefix}eion'],bv[f'{prefix}A'],bv[f'{prefix}B'],
                                        bv[f'{prefix}C'],bv[f'{prefix}D'],bv[f'{prefix}E']])))
        plt.plot(xs, eadata, '--', label=prefix, linewidth=1)
      ### plot urdam components
      for i, ciparams in enumerate(params['ci']):
        cidata = younger(xs, **dict(zip(['eion','A','B','C','D','E'], ciparams[1:])))
        plt.plot(xs, cidata, '-.', label=f'ciUrd{i}')
      for i, eaparams in enumerate(params['ea']):
        eadata = mewe(xs, **dict(zip(['eion','A','B','C','D','E'], eaparams[1:])))
        plt.plot(xs, eadata, '-.', label=f'eaUrd{i}')
      
      plt.legend()
      plt.yscale('log')
      plt.xscale('log')
    #print(result.fit_report())
    #paramsDict[mcInd] = result.best_values
    paramsDict[mcInd] = bv
  # ipdb.set_trace()
  # ipdb.pm()
  ### compare results to accepted rates
  Tlist = numpy.logspace(lowTempPower, highTempPower, numTempSteps) # some temperatures
  CIratesExp = []
  EAratesExp = []
  totalExpRates = []
  for mcInd in range(1,monteCarloLength+1):
    rates = np.zeros(len(Tlist))
    for i, CIprefix in enumerate(CIprefixes):
      vals = paramsDict[mcInd]
      eion = vals[f'{CIprefix}eion']
      A = vals[f'{CIprefix}A']
      B = vals[f'{CIprefix}B']
      C = vals[f'{CIprefix}C']
      D = vals[f'{CIprefix}D']
      E = vals[f'{CIprefix}E']
      tmpRates = calc_ci_urdam(Tlist,[i,eion, A,B,C,D,E])
      rates+=tmpRates
    CIratesExp.append(rates)

    rates = np.zeros(len(Tlist))
    for i, EAprefix in enumerate(EAprefixes):
      vals = paramsDict[mcInd]
      eion = vals[f'{EAprefix}eion']
      A = vals[f'{EAprefix}A']
      B = vals[f'{EAprefix}B']
      C = vals[f'{EAprefix}C']
      D = vals[f'{EAprefix}D']
      E = vals[f'{EAprefix}E']
      tmpRates = calc_ea_urdam(Tlist,[i,eion, A,B,C,D,E])
      rates+=tmpRates
    EAratesExp.append(rates)
    totalExpRates.append(CIratesExp[-1]+EAratesExp[-1])

  CIratesUrd, EAratesUrd = get_ionrec_rate(Z, z1, elsymb, Tlist)
  totalUrdRates = np.sum([CIratesUrd, EAratesUrd], axis=0)

  #remove rate curves with negative values
  expRatesNonNeg = []
  for rate in totalExpRates:
    if sum(rate<0)==0: #if there are no negative values in this rate curve
      expRatesNonNeg.append(rate)
  avgExpRates = np.mean(expRatesNonNeg, axis=0)

  upperExpConf = np.percentile(expRatesNonNeg, 84, axis=0)
  lowerExpConf = np.percentile(expRatesNonNeg, 16, axis=0)
  avgExpRatesUnc = [abs(avgExpRates-lowerExpConf), abs(avgExpRates-upperExpConf)] #[lower, upper] uncertainties

  if makePlotsRates:
    fig, (ax1, ax2) = plt.subplots(2,1)

    cmap = plt.get_cmap("hsv", monteCarloLength)
    for j, expRates in enumerate(expRatesNonNeg):
      ax1.plot(Tlist, expRates,label=f'MC {j}', linewidth=0.75, alpha=0.75, color=cmap(j))
    ax1.plot(Tlist, totalUrdRates, '--', label='urdam')
    ax1.errorbar(Tlist, avgExpRates, avgExpRatesUnc, label='Mean', linewidth=2, color='r')
    #ax1.fill_between(Tlist, avgExpRates-avgExpRatesUnc, avgExpRates+avgExpRatesUnc, color='r', alpha=0.15)
    ax1.set_yscale('log')
    ax1.set_xscale('log')
    ax1.set_title(f'{len(expRatesNonNeg)} Ionization rates of {elementSymbol}{finalChargeState-1}+ to {elementSymbol}{finalChargeState}+ (excluded {len(totalExpRates)-len(expRatesNonNeg)})')
    ax1.set_xlabel('Temperature (K)')
    ax1.set_ylabel('Rate (1/s)')
    ax1.set_ylim(1e-13,1e-7)

    ax2.plot(Tlist, avgExpRates/totalUrdRates, label='Exp/Urd')
    ax2.plot(Tlist, (avgExpRates+avgExpRatesUnc[1])/totalUrdRates, '--', color='g',label='Exp/Urd')
    ax2.plot(Tlist, (avgExpRates-avgExpRatesUnc[0])/totalUrdRates, '--', color='g',label='Exp/Urd')
    ax2.set_xscale('log')
    ax2.set_title('Experiment/Urdampilleta')
    #plt.legend()
    print(f'For {elementSymbol}{finalChargeState-1}+ to {elementSymbol}{finalChargeState}+:\n\tExcluded {len(totalExpRates)-len(expRatesNonNeg)} of {len(totalExpRates)} curves due to negative values.')
    #plt.savefig(f'{elementSymbol}{finalChargeState-1}+ to {elementSymbol}{finalChargeState}+.pdf')
  return Tlist, avgExpRates, avgExpRatesUnc

def getIonrecArgsFromFile(element: str):
  argsDict={}
  searchstringexp = os.path.join(os.path.dirname(__file__), 'ionrecFitSettings.xlsx')
  fname = glob.glob(searchstringexp)[0]
  if fname:
    #load function args per charge state
    wb = openpyxl.load_workbook(fname)
    ws = wb[element]#wb.active
    ws_contents=list(ws.iter_cols(values_only=True))
    for arg in ws_contents:
      argsDict[arg[0]]=[val for val in arg[1:] if val is not None]
  else:
    print("Failed to find settings file. Place in same directory as make_ionrec.py")
  return argsDict

if __name__ == '__main__':
  import matplotlib.pyplot as plt
  plt.ion()
  if False:
    ionrecAnalysis(Z=8, finalChargeState=7, elementSymbol='O', monteCarloLength=6, numCIModels=1, numEAModels=0, 
                    lowTempPower=4, highTempPower=9, numTempSteps=300, makePlots=True, makePlotsRates=True, FixD=True, addZeroPt=True)
  if True:
    element = 'O'
    ratesAndUncs = {}
    ionrecArgs = getIonrecArgsFromFile(element)
    for i, chargeState in enumerate(ionrecArgs['finalChargeState']):
      tmpArgsDict = {}
      for key in ionrecArgs.keys():
        tmpArgsDict[key] = ionrecArgs[key][i]
      Tlist, rates, uncs = ionrecAnalysis(**tmpArgsDict)
      ratesAndUncs[f'{element}{chargeState-1}+']={'Tlist':Tlist, 'rates':rates, 'uncsLower':uncs[0], 'uncsUpper':uncs[1]}
    import matplotlib.backends.backend_pdf
    pdf = matplotlib.backends.backend_pdf.PdfPages("OxygenOutput.pdf")
    for fig in range(1, plt.gcf().number + 1): ## will open an empty extra figure :(
        pdf.savefig( fig )
    pdf.close()
    picklename = f'{element}.pickle'
    if os.path.exists(os.path.abspath(os.path.join('pickle',picklename))):
      fexists = input("Pickle file already exists. Overwrite? [y]/[n]:\t")
      if fexists=='y':
        with open(os.path.abspath(os.path.join('pickle',picklename)), 'wb') as file:
          pickle.dump(ratesAndUncs, file)
          print(f"Saved pickle file to {os.path.abspath(os.path.join('pickle',picklename))}")
      else:
        print("Pickle file already exists; did not overwrite.")

